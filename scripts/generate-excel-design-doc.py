#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
画面設計書Excel生成スクリプト
"""

import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# パス設定
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
SCREENSHOT_DIR = os.path.join(PROJECT_ROOT, 'docs', 'screenshots')
OUTPUT_PATH = os.path.join(PROJECT_ROOT, 'docs', '画面設計書.xlsx')

# 画面定義
SCREENS = [
    {
        'id': '01',
        'name': 'メイン画面',
        'description': '医療機器管理システムのメイン画面。資産検索条件を入力し、検索を実行する。',
        'features': [
            '施設情報検索（施設名、病床規模、部門名、部署名）',
            '資産情報検索（Category、大分類、中分類、品名、メーカー、型式）',
            '検索実行と結果表示',
            '個体管理リスト作成',
            'マスタ管理',
            'QRコード管理',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'システムロゴ（SHIP HEALTHCARE）、個体管理リスト作成ボタン、マスタ管理ボタン、QRコード管理ボタン、ログアウトボタン'},
            {'name': '検索条件読み込みエリア', 'description': '保存済み検索条件をドロップダウンから選択して読み込む機能'},
            {'name': '施設情報セクション', 'description': '施設名、病床規模（範囲指定）、部門名、部署名を入力するフォーム'},
            {'name': '資産情報セクション', 'description': 'Category、大分類、中分類、品名、メーカー、型式を入力するフォーム'},
            {'name': '検索ボタン', 'description': '入力された条件で資産を検索実行'},
        ]
    },
    {
        'id': '02',
        'name': 'QRコード発行画面',
        'description': '資産管理用のQRコードを新規発行または再発行する画面。',
        'features': [
            'QRコード番号入力（アルファベット-2桁数字-5桁数字）',
            '新規発行/再発行のタブ切り替え',
            'ラベルテンプレート選択（QRコード：小/中/大、バーコード：小/中/大）',
            'フッター記入項目',
            '発行枚数指定（最大100枚）',
            '発行予定番号範囲のプレビュー',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': '戻るボタン、画面タイトル（QRコード新規発行）'},
            {'name': 'タブ切り替え', 'description': '新規発行/再発行の切り替えタブ'},
            {'name': 'QRコード番号入力', 'description': 'アルファベット（A-Z）、2桁数字、5桁数字（開始番号）を入力'},
            {'name': 'ラベルテンプレート選択', 'description': 'QRコード（小18mm/中24mm/大36mm）、バーコード（小18mm/中24mm/大36mm）から選択'},
            {'name': 'フッター記入項目', 'description': 'ラベルのフッターに印字するテキスト（文字数制限あり）'},
            {'name': '発行枚数', 'description': '発行する枚数を入力（最大100枚）'},
            {'name': '発行予定番号範囲', 'description': '開始番号から終了番号までの範囲と枚数をプレビュー表示'},
            {'name': '印刷プレビューへボタン', 'description': '次の印刷プレビュー画面へ遷移'},
        ]
    },
    {
        'id': '03',
        'name': 'QRコード印刷プレビュー画面',
        'description': 'QRコードの印刷前にプレビューを確認する画面。',
        'features': [
            'QRコード一覧表示',
            '印刷プレビュー',
            '印刷実行',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': '戻るボタン、画面タイトル（QRコード印刷プレビュー）'},
            {'name': 'QRコード一覧', 'description': '発行予定のQRコードを一覧で表示（番号、QRコード画像、フッターテキスト）'},
            {'name': '印刷実行ボタン', 'description': 'ラベルプリンターで印刷を実行'},
            {'name': 'キャンセルボタン', 'description': '印刷をキャンセルして前画面に戻る'},
        ]
    },
    {
        'id': '04',
        'name': '資産検索結果画面',
        'description': 'メイン画面での検索結果を一覧表示する画面。',
        'features': [
            '検索結果一覧表示',
            '検索条件フィルター',
            '表示列カスタマイズ',
            '資産詳細表示',
            'データエクスポート',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、検索結果件数、メニュー（申請一覧/見積書管理）、戻るボタン'},
            {'name': 'フィルターヘッダー', 'description': '施設名、部門名、Category、大分類、中分類、品名、メーカー、型式でのフィルター'},
            {'name': '表示列選択', 'description': '表示する列を選択（施設名、部門名等）'},
            {'name': '検索結果テーブル', 'description': '資産情報を表形式で表示（横スクロール対応）'},
            {'name': 'アクションボタン', 'description': '詳細表示、編集、削除等の操作ボタン'},
        ]
    },
    {
        'id': '05',
        'name': '現有資産調査画面',
        'description': '現有資産の分類情報を登録・管理する画面。',
        'features': [
            '分類情報入力（大分類、中分類、品名、メーカー、型式）',
            'QRコードスキャン',
            '登録履歴表示',
            '一時保存機能',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': '戻るボタン、画面タイトル（現有資産調査）'},
            {'name': 'QRコードスキャンボタン', 'description': 'QRコードをスキャンして資産を特定'},
            {'name': '分類情報入力フォーム', 'description': '大分類、中分類、品名、メーカー、型式を選択式で入力（Choices.js使用）'},
            {'name': '登録履歴ボタン', 'description': '過去の登録履歴一覧を表示'},
            {'name': '一時保存ボタン', 'description': '入力内容を一時保存'},
            {'name': '登録ボタン', 'description': '入力内容を本登録'},
            {'name': 'フッター', 'description': '操作ボタン群（保存、登録、履歴表示）'},
        ]
    },
    {
        'id': '06',
        'name': '登録履歴画面',
        'description': '過去に登録した資産情報の履歴を確認・再利用する画面。',
        'features': [
            '登録履歴一覧（カード形式）',
            '履歴の選択',
            'カード内容の編集',
            '履歴データの再利用',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': '戻るボタン、画面タイトル（登録履歴）、件数表示'},
            {'name': '履歴カード', 'description': '過去の登録情報をカード形式で表示（大分類、中分類、品名、メーカー、型式、登録日時）'},
            {'name': 'カード選択チェックボックス', 'description': '履歴カードを選択'},
            {'name': 'フッター', 'description': '修正ボタン（インライン編集）、再利用ボタン（前画面に情報をセット）'},
        ]
    },
    {
        'id': '07',
        'name': '申請一覧画面',
        'description': '資産に関する各種申請を一覧表示し、管理する画面。',
        'features': [
            '申請一覧表示',
            '申請種別フィルター（新規購入、増設購入、更新購入、移動、廃棄、保留）',
            '状態フィルター（下書き、承認待ち、承認済み、差し戻し、却下）',
            '見積依頼Noでの検索',
            '申請詳細表示',
            '見積グルーピング機能',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、申請一覧タイトル、件数表示、メニュー（申請一覧/見積書管理）、戻るボタン'},
            {'name': 'フィルターヘッダー', 'description': '申請種別、状態、見積依頼No、申請日（開始/終了）、キーワード検索、クリアボタン'},
            {'name': '申請一覧テーブル', 'description': '申請番号、申請日、申請種別、資産情報、数量、見積依頼No、購入先店舗、見積情報、状態、承認進捗、アクション'},
            {'name': 'チェックボックス', 'description': '複数申請を選択して一括操作'},
            {'name': '一括操作バー', 'description': '選択した申請をまとめて見積グルーピング'},
            {'name': 'アクションボタン', 'description': '詳細表示、編集、削除等の操作'},
        ]
    },
    {
        'id': '08',
        'name': '見積依頼一覧画面',
        'description': '見積依頼を一覧表示し、管理する画面。',
        'features': [
            '見積依頼一覧表示',
            '購入先店舗でのフィルター',
            '状態フィルター',
            '見積依頼詳細表示',
            '見積書アップロード',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、見積依頼一覧タイトル、件数表示、メニュー、戻るボタン'},
            {'name': 'フィルターヘッダー', 'description': '購入先店舗、状態、作成日（開始/終了）、キーワード検索'},
            {'name': '見積依頼一覧テーブル', 'description': '見積依頼No、作成日、購入先店舗、申請件数、総額、見積書、状態、アクション'},
            {'name': 'アクションボタン', 'description': '詳細表示、見積書アップロード、編集、削除'},
        ]
    },
    {
        'id': '09',
        'name': '見積DataBox画面',
        'description': 'アップロードされた見積書を管理する画面。',
        'features': [
            '見積書一覧表示',
            '見積書プレビュー',
            '見積書ダウンロード',
            '見積書削除',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、見積DataBoxタイトル、件数表示、メニュー、戻るボタン'},
            {'name': 'フィルターヘッダー', 'description': '見積依頼No、購入先店舗、アップロード日（開始/終了）、キーワード検索'},
            {'name': '見積書一覧', 'description': '見積依頼No、購入先店舗、ファイル名、アップロード日、ファイルサイズ、アクション'},
            {'name': 'アクションボタン', 'description': 'プレビュー、ダウンロード、削除'},
        ]
    },
    {
        'id': '10',
        'name': '個体管理リスト画面',
        'description': '個体ごとの資産管理リストを表示する画面。',
        'features': [
            '個体管理リスト一覧',
            'QRコード生成',
            'リストのエクスポート',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、個体管理リストタイトル、件数表示、戻るボタン'},
            {'name': 'フィルターヘッダー', 'description': '施設名、資産種別、QRコード有無、キーワード検索'},
            {'name': '個体リスト一覧', 'description': '資産番号、資産名、メーカー、型式、QRコード、アクション'},
            {'name': 'QRコード生成ボタン', 'description': '選択した資産のQRコードを生成'},
            {'name': 'エクスポートボタン', 'description': 'リストをExcel/CSV形式でエクスポート'},
        ]
    },
    {
        'id': '11',
        'name': 'SHIP施設マスタ画面',
        'description': 'SHIP連携用の施設マスタを管理する画面。',
        'features': [
            '施設マスタ一覧表示',
            '施設情報の編集',
            '施設情報の削除',
            '新規施設追加',
            'フィルター機能',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、SHIP施設マスタタイトル、件数表示、戻るボタン'},
            {'name': 'フィルタートグル', 'description': 'フィルター表示/非表示の切り替え'},
            {'name': 'フィルターパネル', 'description': '施設コード、施設名、都道府県、病床規模での検索'},
            {'name': '操作ボタンバー', 'description': '新規登録、Excel出力、一括削除ボタン'},
            {'name': '施設マスタテーブル', 'description': '施設コード、施設名、都道府県、病床規模、部門数、最終更新日、アクション'},
            {'name': 'チェックボックス', 'description': '複数施設を選択して一括操作'},
            {'name': 'アクションボタン', 'description': '詳細表示、編集、削除'},
            {'name': 'ページネーション', 'description': 'ページ切り替え（前へ/次へ）'},
        ]
    },
    {
        'id': '12',
        'name': 'SHIP資産マスタ画面',
        'description': 'SHIP連携用の資産マスタを管理する画面。',
        'features': [
            '資産マスタ一覧表示',
            '資産情報の編集',
            '資産情報の削除',
            '新規資産追加',
            'フィルター機能',
        ],
        'elements': [
            {'name': 'ヘッダー', 'description': 'SHIPロゴ、SHIP資産マスタタイトル、件数表示、戻るボタン'},
            {'name': 'フィルタートグル', 'description': 'フィルター表示/非表示の切り替え'},
            {'name': 'フィルターパネル', 'description': '資産コード、資産名、Category、大分類、中分類での検索'},
            {'name': '操作ボタンバー', 'description': '新規登録、Excel出力、一括削除ボタン'},
            {'name': '資産マスタテーブル', 'description': '資産コード、資産名、Category、大分類、中分類、品名、メーカー、型式、最終更新日、アクション'},
            {'name': 'チェックボックス', 'description': '複数資産を選択して一括操作'},
            {'name': 'アクションボタン', 'description': '詳細表示、編集、削除'},
            {'name': 'ページネーション', 'description': 'ページ切り替え（前へ/次へ）'},
        ]
    }
]

# スタイル定義
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
SECTION_FONT = Font(name='メイリオ', size=11, bold=True)
NORMAL_FONT = Font(name='メイリオ', size=10)
TITLE_FONT = Font(name='メイリオ', size=14, bold=True)
ELEMENT_HEADER_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_index_sheet(wb):
    """目次シートを作成"""
    ws = wb.active
    ws.title = '目次'

    # タイトル
    ws['A1'] = '医療機器管理システム 画面設計書'
    ws['A1'].font = Font(name='メイリオ', size=16, bold=True)
    ws.merge_cells('A1:D1')

    # 作成日
    ws['A2'] = f'作成日: {datetime.now().strftime("%Y年%m月%d日")}'
    ws['A2'].font = NORMAL_FONT
    ws.merge_cells('A2:D2')

    # ヘッダー
    row = 4
    ws[f'A{row}'] = 'No.'
    ws[f'B{row}'] = '画面ID'
    ws[f'C{row}'] = '画面名'
    ws[f'D{row}'] = '説明'

    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 画面一覧
    row += 1
    for idx, screen in enumerate(SCREENS, 1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = screen['id']
        ws[f'C{row}'] = screen['name']
        ws[f'D{row}'] = screen['description']

        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col in ['A', 'B']:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        row += 1

    # 列幅調整
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 60


def create_screen_sheet(wb, screen):
    """各画面のシートを作成"""
    ws = wb.create_sheet(title=f"{screen['id']}_{screen['name']}")

    row = 1

    # タイトル
    ws[f'A{row}'] = f"{screen['id']}. {screen['name']}"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:F{row}')
    row += 2

    # 概要
    ws[f'A{row}'] = '画面概要'
    ws[f'A{row}'].font = SECTION_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    ws[f'A{row}'] = screen['description']
    ws[f'A{row}'].font = NORMAL_FONT
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 40
    row += 2

    # 主な機能
    ws[f'A{row}'] = '主な機能'
    ws[f'A{row}'].font = SECTION_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    for feature in screen['features']:
        ws[f'A{row}'] = f'• {feature}'
        ws[f'A{row}'].font = NORMAL_FONT
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    row += 1

    # 画面要素
    if 'elements' in screen and screen['elements']:
        ws[f'A{row}'] = '画面要素'
        ws[f'A{row}'].font = SECTION_FONT
        ws[f'A{row}'].fill = SECTION_FILL
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # テーブルヘッダー
        ws[f'A{row}'] = '要素名'
        ws[f'B{row}'] = '説明'
        for col in ['A', 'B']:
            cell = ws[f'{col}{row}']
            cell.font = Font(name='メイリオ', size=10, bold=True)
            cell.fill = ELEMENT_HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        ws.merge_cells(f'B{row}:F{row}')
        row += 1

        # 要素一覧
        for element in screen['elements']:
            ws[f'A{row}'] = element['name']
            ws[f'B{row}'] = element['description']

            cell_a = ws[f'A{row}']
            cell_a.font = NORMAL_FONT
            cell_a.border = THIN_BORDER
            cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            cell_b = ws[f'B{row}']
            cell_b.font = NORMAL_FONT
            cell_b.border = THIN_BORDER
            cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            ws.merge_cells(f'B{row}:F{row}')
            ws.row_dimensions[row].height = 30
            row += 1

        row += 1

    # スクリーンショット
    ws[f'A{row}'] = 'スクリーンショット'
    ws[f'A{row}'].font = SECTION_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    # デバイスごとのスクリーンショット
    devices = [
        ('desktop', 'デスクトップ (1920x1080)', 70),
        ('tablet', 'タブレット (768x1024)', 50),
        ('mobile', 'モバイル (375x667)', 35)
    ]

    for device, label, height in devices:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='メイリオ', size=10, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row}'].fill = ELEMENT_HEADER_FILL
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # 画像挿入
        img_path = os.path.join(SCREENSHOT_DIR, f"{screen['id']}_{screen['name']}_{device}.png")
        if os.path.exists(img_path):
            try:
                img = Image(img_path)
                # 画像サイズを調整（高さを基準に）
                aspect_ratio = img.width / img.height
                img.height = height * 15  # セル高さに合わせる（ポイント → ピクセル変換）
                img.width = img.height * aspect_ratio

                ws.add_image(img, f'A{row}')
                # 画像用の行の高さを設定
                ws.row_dimensions[row].height = height * 15 / 1.33  # ピクセル → ポイント変換
            except Exception as e:
                ws[f'A{row}'] = f'画像読み込みエラー: {str(e)}'
                ws[f'A{row}'].font = NORMAL_FONT
        else:
            ws[f'A{row}'] = '画像が見つかりません'
            ws[f'A{row}'].font = NORMAL_FONT

        row += 1
        row += 1  # 余白

    # 列幅調整
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15


def generate_excel():
    """Excelファイルを生成"""
    print('Excel画面設計書の生成を開始します...')

    # ワークブック作成
    wb = Workbook()

    # 目次シート作成
    print('目次シートを作成中...')
    create_index_sheet(wb)

    # 各画面のシート作成
    for screen in SCREENS:
        print(f'{screen["id"]}. {screen["name"]} のシートを作成中...')
        create_screen_sheet(wb, screen)

    # 保存
    print(f'\n保存中: {OUTPUT_PATH}')
    wb.save(OUTPUT_PATH)
    print('✓ Excel画面設計書の生成が完了しました')


if __name__ == '__main__':
    generate_excel()
